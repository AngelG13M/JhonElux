import pandas as pd
import streamlit as st
import os
import shutil 
import json 
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font
from io import BytesIO
from PIL import Image
from PIL.ExifTags import TAGS
import gc # Mantenemos el Garbage Collector para limpieza de RAM

# ----------------------------------------------------
# CONFIGURACIÓN DINÁMICA Y PERSISTENCIA
# ----------------------------------------------------
CONFIG_FILE = 'config_cols.json'
PERSISTENCE_FILE = 'datos_maestro.csv' 
IMAGE_FOLDER = 'imagenes_persistentes' 

# Función para cargar la configuración de columnas
def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f: 
                return json.load(f)
        except Exception as e:
            st.error(f"Error al leer config_cols.json (usando UTF-8). Por favor, verifique el formato del archivo. Error: {e}")
            return {
                "CONDICIONES_INSPECCION": ["DAÑO EN EMPAQUE", "DAÑO FISICO", "ACCESORIOS COMPLETOS", "PARILLA EN MAL ESTADO", "PRESENTA RESTOS METALICOS (VIRUTAS)", "TAPAS PRESENTAN OXIDO", "PRESENTA RAYAS", "TARJETA DE GARANTÍA", "TIENE ETIQUETA DE EFICIENCIA ENERGETICA"],
                "COLUMNAS_IMAGEN": ["FOTO DE SERIE", "FOTO DEL EMPAQUE", "FOTO DE PRODUCTO COMPLETO", "FOTO PARTE TRASERA", "FOTO DE OBSERVACIONES A 50 CM (VIRUTAS)", "FOTO DE OBSERVACIONES CERCA (VIRUTAS)", "FOTO DE OBSERVACIONES A 50 CM (OXIDO EN TAPILLAS)", "FOTO DE OBSERVACIONES CERCA (OXIDO EN TAPILLAS)", "FOTO DE OBSERVACIONES A 50 CM (MANCHAS)", "FOTO DE OBSERVACIONES CERCA (MANCHAS)", "FOTO DE OBSERVACIONES A 50 CM (RAYAS)", "FOTO DE OBSERVACIONES CERCA (RAYAS)", "FOTO DE ACCESORIOS"]
            }
    return {
        "CONDICIONES_INSPECCION": ["DAÑO EN EMPAQUE", "DAÑO FISICO", "ACCESORIOS COMPLETOS", "PARILLA EN MAL ESTADO", "PRESENTA RESTOS METALICOS (VIRUTAS)", "TAPAS PRESENTAN OXIDO", "PRESENTA RAYAS", "TARJETA DE GARANTÍA", "TIENE ETIQUETA DE EFICIENCIA ENERGETICA"],
        "COLUMNAS_IMAGEN": ["FOTO DE SERIE", "FOTO DEL EMPAQUE", "FOTO DE PRODUCTO COMPLETO", "FOTO PARTE TRASERA", "FOTO DE OBSERVACIONES A 50 CM (VIRUTAS)", "FOTO DE OBSERVACIONES CERCA (VIRUTAS)", "FOTO DE OBSERVACIONES A 50 CM (OXIDO EN TAPILLAS)", "FOTO DE OBSERVACIONES CERCA (OXIDO EN TAPILLAS)", "FOTO DE OBSERVACIONES A 50 CM (MANCHAS)", "FOTO DE OBSERVACIONES CERCA (MANCHAS)", "FOTO DE OBSERVACIONES A 50 CM (RAYAS)", "FOTO DE OBSERVACIONES CERCA (RAYAS)", "FOTO DE ACCESORIOS"]
    }

# Cargar la configuración y definir las listas dinámicas
config = load_config()
CONDICIONES_INSPECCION = config["CONDICIONES_INSPECCION"]
COLUMNAS_IMAGEN = config["COLUMNAS_IMAGEN"]

# DEFINICIÓN DINÁMICA DE ENCABEZADOS
ENCABEZADOS = ['MODELO', 'SERIE']
ENCABEZADOS.extend(CONDICIONES_INSPECCION) 
ENCABEZADOS.append('OBSERVACIONES') 

COLUMNAS_FINALES = ENCABEZADOS + COLUMNAS_IMAGEN

# --- 2. Constantes de Formato para Excel ---
ALTURA_ENCABEZADO_PT = 60  
ALTURA_FILA_DATOS_PT = 75  
ANCHO_COLUMNA_NORMAL_UNITS = 14 
ANCHO_COLUMNA_OBSERVACIONES_UNITS = 28 
IMAGEN_WIDTH = 150 # Restauramos un tamaño decente para alta calidad
IMAGEN_HEIGHT = 150 # Restauramos un tamaño decente para alta calidad
ALINEACION_CENTRO = Alignment(horizontal='center', vertical='center', wrap_text=True)
RELLENO_VERDE_AZULADO = PatternFill(start_color='20B2AA', end_color='20B2AA', fill_type='solid') 
FUENTE_ENCABEZADO = Font(color='FFFFFF', bold=True) 

# --- 3. Funciones de Lógica y Persistencia ---

def cargar_datos_persistentes():
    if os.path.exists(PERSISTENCE_FILE):
        try:
            df = pd.read_csv(PERSISTENCE_FILE, dtype=str) 
            
            for col in COLUMNAS_IMAGEN:
                if col not in df.columns:
                    df[col] = ''
            
            for col in ENCABEZADOS:
                 if col not in df.columns:
                    df[col] = ''
            
            return df.to_dict('records')
        except Exception as e:
            st.error(f"Error al cargar datos persistentes: {e}. Se inicia una lista vacía.")
            return []
    return []

def generar_excel_con_formato(df):
    """Genera el archivo Excel en memoria, leyendo las imágenes desde el disco con MÁXIMA CALIDAD."""
    
    os.makedirs(IMAGE_FOLDER, exist_ok=True) 

    output = BytesIO()
    df_datos_texto = df[[c for c in ENCABEZADOS if c in df.columns]] 
    df_datos_texto.to_excel(output, index=False)
    output.seek(0)
    
    wb = load_workbook(output)
    ws = wb.active 
    
    ws.row_dimensions[1].height = ALTURA_ENCABEZADO_PT
    
    for row_index in range(1, len(df) + 2):
        fila_excel_obj = ws[row_index]
        if row_index > 1:
            ws.row_dimensions[row_index].height = ALTURA_FILA_DATOS_PT
            
        for col_idx_0based in range(len(ENCABEZADOS)): 
            celda = fila_excel_obj[col_idx_0based]
            celda.alignment = ALINEACION_CENTRO 
            
            if row_index == 1:
                celda.fill = RELLENO_VERDE_AZULADO
                celda.font = FUENTE_ENCABEZADO
        
    for idx, col_name in enumerate(ENCABEZADOS, start=1):
        col_letter = get_column_letter(idx)
        if col_name == 'OBSERVACIONES':
            ws.column_dimensions[col_letter].width = ANCHO_COLUMNA_OBSERVACIONES_UNITS
        else:
            ws.column_dimensions[col_letter].width = ANCHO_COLUMNA_NORMAL_UNITS

    # Inserción de Imágenes Incrustadas
    df_rutas_imagen = df[COLUMNAS_IMAGEN]
    for indice_fila, (index, fila_rutas) in enumerate(df_rutas_imagen.iterrows()):
        fila_excel = indice_fila + 2
        
        for indice_columna, nombre_columna_foto in enumerate(COLUMNAS_IMAGEN):
            
            imagen_ruta_guardada = fila_rutas[nombre_columna_foto]
            path_check = str(imagen_ruta_guardada).strip()
            es_path_valido = path_check and path_check.lower() != 'none' and path_check.lower() != 'nan'
            
            columna_imagen_index = len(ENCABEZADOS) + indice_columna + 1
            columna_letra = get_column_letter(columna_imagen_index)

            if indice_fila == 0: 
                 celda_encabezado = ws[f'{columna_letra}1']
                 celda_encabezado.value = nombre_columna_foto
                 celda_encabezado.alignment = ALINEACION_CENTRO 
                 celda_encabezado.fill = RELLENO_VERDE_AZULADO
                 celda_encabezado.font = FUENTE_ENCABEZADO
                 ws.column_dimensions[columna_letra].width = ANCHO_COLUMNA_NORMAL_UNITS
            
            if es_path_valido and os.path.exists(path_check):
                try:
                    with open(path_check, 'rb') as f:
                        img_pil = Image.open(f)
                        img_pil.load() 
                        
                    exif = img_pil._getexif()
                    orientation = 1
                    if exif:
                        for tag, value in exif.items():
                            if TAGS.get(tag, tag) == 'Orientation':
                                orientation = value
                                break
                    
                    if orientation == 3: img_pil = img_pil.rotate(180, expand=True)
                    elif orientation == 6: img_pil = img_pil.rotate(-90, expand=True)
                    elif orientation == 8: img_pil = img_pil.rotate(90, expand=True)

                    # No comprimimos, solo rotamos y guardamos el resultado
                    rotated_img_bytes = BytesIO()
                    img_pil.save(rotated_img_bytes, format='PNG') 
                    rotated_img_bytes.seek(0)

                    # Insertar la imagen de alta calidad
                    img = OpenpyxlImage(rotated_img_bytes)
                    img.width = IMAGEN_WIDTH 
                    img.height = IMAGEN_HEIGHT
                    ws.add_image(img, f'{columna_letra}{fila_excel}')
                    
                    # Limpieza de Memoria
                    del img_pil
                    del rotated_img_bytes
                    gc.collect() 

                except Exception as e:
                    st.error(f"❌ Error crítico al insertar imagen: {e}")
                    # Ya que la calidad es importante, no incluimos la opción de baja calidad aquí

    final_output = BytesIO()
    wb.save(final_output)
    return final_output


def guardar_registro_y_limpiar(modelo, serie, condiciones, observaciones, fotos):
    """Guarda el registro, guarda las imágenes físicamente y actualiza el archivo persistente."""
    
    if not modelo or not serie:
        st.error("❌ Los campos MODELO y SERIE son obligatorios.")
        return 

    os.makedirs(IMAGE_FOLDER, exist_ok=True)
    
    # Guardar las fotos físicamente y obtener la RUTA
    fotos_rutas_guardadas = {}
    
    for k, uploaded_file in fotos.items():
        if uploaded_file is not None:
            file_extension = uploaded_file.name.split('.')[-1]
            timestamp = pd.Timestamp.now().strftime("%Y%m%d%H%M%S")
            filename = f"{modelo.replace(' ', '_')}_{serie.replace(' ', '_')}_{k.replace(' ', '_')}_{timestamp}.{file_extension}"
            ruta_guardado = os.path.join(IMAGE_FOLDER, filename)
            
            with open(ruta_guardado, "wb") as f:
                f.write(uploaded_file.getbuffer()) 
                
            fotos_rutas_guardadas[k] = ruta_guardado
        else:
            fotos_rutas_guardadas[k] = None

    # Preparar el nuevo registro
    nuevo_registro = {
        'MODELO': modelo,
        'SERIE': serie,
        'OBSERVACIONES': observaciones,
        **{k: 'SÍ' if v else 'NO' for k, v in condiciones.items()},
        **fotos_rutas_guardadas 
    }
    
    st.session_state['datos_ingresados'].append(nuevo_registro)

    # Persistencia: Guardar el registro a un archivo CSV (texto + rutas)
    try:
        df_actual = pd.DataFrame(st.session_state['datos_ingresados'])
        df_actual[[c for c in COLUMNAS_FINALES if c in df_actual.columns]].to_csv(PERSISTENCE_FILE, index=False)
    except Exception as e:
        st.warning(f"Advertencia: No se pudo guardar el archivo de persistencia CSV. Error: {e}")
        
    st.success(f"✅ Registro para Modelo {modelo} añadido a la lista. Ingresa el siguiente.")
    
    st.session_state['limpiador_key'] += 1 
    
    keys_a_limpiar_texto = ['input_modelo', 'input_serie', 'input_observaciones']
    for key in keys_a_limpiar_texto:
        if key in st.session_state:
            st.session_state[key] = ""
            
    st.rerun()

# ----------------------------------------------------
# PUNTO DE ENTRADA Y DISEÑO DE LA INTERFAZ PRINCIPAL
# ----------------------------------------------------
def main():
    st.set_page_config(page_title="Electrolux Inspección", layout="wide")

    LOGO_PATH = "electrolux_logo.png"

    # APLICACIÓN DE LOGO Y TÍTULO EN LA BARRA SUPERIOR
    try:
        if os.path.exists(LOGO_PATH):
            col_logo, col_titulo = st.columns([1.5, 4]) 
            
            with col_logo:
                 st.image(LOGO_PATH, width=150) 
            
            col_titulo.title("Generador de Reporte de Inspección")
            
        else:
            st.title("📋 Generador de Reporte de Inspección")
            st.warning(f"⚠️ ¡Falta el logo! Coloque '{LOGO_PATH}' en esta carpeta.")
    except Exception:
        st.title("📋 Generador de Reporte de Inspección")
    
    # INSERCIÓN DEL BOTÓN DE ACCESO A ADMINISTRACIÓN
    st.markdown("---") 
    
    if st.button("⚙️ Editar Formato de Columnas (Administración)", type="secondary"):
        st.write("<meta http-equiv='refresh' content='0; url=admin_columnas'>", unsafe_allow_html=True)
        
    st.markdown("---")
    
    # --- RESTO DE LA LÓGICA DE LA PÁGINA ---
    if 'datos_ingresados' not in st.session_state:
        st.session_state['datos_ingresados'] = cargar_datos_persistentes()
    if 'limpiador_key' not in st.session_state:
        st.session_state['limpiador_key'] = 0

    # Usamos st.form para agrupar los inputs
    with st.form(key='inspeccion_form'):
        
        st.markdown("<h3 style='color: #FFFFFF;'>--- Datos de Producto y Condición ---</h3>", unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        
        # Modelo y Serie
        modelo = col1.text_input("MODELO", key='input_modelo')
        serie = col2.text_input("SERIE", key='input_serie')
        
        # Checkboxes
        condiciones = {}
        st.markdown("##### Condiciones de Inspección")
        cols = st.columns(4) 
        for i, col in enumerate(CONDICIONES_INSPECCION):
            clave_dinamica = f'check_{col}_{st.session_state["limpiador_key"]}'
            initial_value = st.session_state.get(clave_dinamica, False)
            condiciones[col] = cols[i % 4].checkbox(col, value=initial_value, key=clave_dinamica)
            
        # Observaciones
        observaciones = st.text_area("OBSERVACIONES", key='input_observaciones')
        
        st.markdown("---")

        # File Uploaders
        st.markdown("<h3 style='color: #FFFFFF;'>📸 Fotos</h3>", unsafe_allow_html=True)
        fotos = {}
        cols_foto = st.columns(4)
        for i, col in enumerate(COLUMNAS_IMAGEN):
            clave_dinamica = f'file_{col}_{st.session_state["limpiador_key"]}'
            fotos[col] = cols_foto[i % 4].file_uploader(col, type=['png', 'jpg', 'jpeg'], key=clave_dinamica)

        st.markdown("---")
        
        # Botón de envío
        st.form_submit_button(
            label="➕ Añadir Registro a la Lista", 
            type="primary", 
            on_click=guardar_registro_y_limpiar,
            args=(modelo, serie, condiciones, observaciones, fotos)
        )

    # --- 5. Mostrar tabla de registros guardados y botón de descarga ---

    if st.session_state['datos_ingresados']:
        df_preview = pd.DataFrame(st.session_state['datos_ingresados'])
        
        st.subheader(f"Registros Guardados ({len(st.session_state['datos_ingresados'])})")
        st.dataframe(df_preview[[c for c in ENCABEZADOS if c in df_preview.columns]], use_container_width=True, height=200)

        # La función de generación se llama SOLO aquí (bajo demanda)
        excel_file = generar_excel_con_formato(df_preview) 
        
        st.download_button(
            label="💾 Descargar Excel Final",
            data=excel_file.getvalue(),
            file_name="Inspeccion_Reporte_Mobil.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        if st.button("🗑️ Limpiar Todos los Registros"):
            st.session_state['datos_ingresados'] = []
            st.session_state['limpiador_key'] += 1 
            
            if os.path.exists(PERSISTENCE_FILE):
                 os.remove(PERSISTENCE_FILE)
            if os.path.exists(IMAGE_FOLDER):
                 shutil.rmtree(IMAGE_FOLDER) 
            
            st.success("Lista de registros y archivos persistentes eliminados.")
            st.rerun()

# Punto de entrada de la aplicación
if __name__ == '__main__':
    main()