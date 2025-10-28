import pandas as pd
import streamlit as st
import os
import shutil 
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font
from io import BytesIO
from PIL import Image
from PIL.ExifTags import TAGS

# ----------------------------------------------------
# ARCHIVOS DE PERSISTENCIA Y CARPETAS
# ----------------------------------------------------
PERSISTENCE_FILE = 'datos_maestro.csv' 
IMAGE_FOLDER = 'imagenes_persistentes' 

# --- 1. Definición de Encabezados ---
ENCABEZADOS = [
    'MODELO', 'SERIE', 'DAÑO EN EMPAQUE', 'DAÑO FISICO', 'ACCESORIOS COMPLETOS', 
    'PARILLA EN MAL ESTADO', 'PRESENTA RESTOS METALICOS (VIRUTAS)', 
    'TAPAS PRESENTAN OXIDO', 'PRESENTA RAYAS', 'TARJETA DE GARANTÍA', 
    'TIENE ETIQUETA DE EFICIENCIA ENERGETICA', 'OBSERVACIONES'
]

COLUMNAS_IMAGEN = [
    'FOTO DE SERIE', 'FOTO DEL EMPAQUE', 'FOTO DE PRODUCTO COMPLETO', 
    'FOTO PARTE TRASERA', 'FOTO DE OBSERVACIONES A 50 CM (VIRUTAS)', 
    'FOTO DE OBSERVACIONES CERCA (VIRUTAS)', 
    'FOTO DE OBSERVACIONES A 50 CM (OXIDO EN TAPILLAS)', 
    'FOTO DE OBSERVACIONES CERCA (OXIDO EN TAPILLAS)', 
    'FOTO DE OBSERVACIONES A 50 CM (MANCHAS)', 
    'FOTO DE OBSERVACIONES CERCA (MANCHAS)', 
    'FOTO DE OBSERVACIONES A 50 CM (RAYAS)', 
    'FOTO DE OBSERVACIONES CERCA (RAYAS)', 'FOTO DE ACCESORIOS'
]

COLUMNAS_FINALES = ENCABEZADOS + COLUMNAS_IMAGEN

# --- 2. Constantes de Formato para Excel ---
ALTURA_FILA_PT = 75 
ANCHO_COLUMNA_NORMAL_UNITS = 14 
ANCHO_COLUMNA_OBSERVACIONES_UNITS = 28 
IMAGEN_WIDTH = 100 
IMAGEN_HEIGHT = 100
ALINEACION_CENTRO = Alignment(horizontal='center', vertical='center', wrap_text=True)
RELLENO_VERDE_AZULADO = PatternFill(start_color='20B2AA', end_color='20B2AA', fill_type='solid') 
FUENTE_ENCABEZADO = Font(color='FFFFFF', bold=True) 

# ----------------------------------------------------
# FUNCIÓN PARA CARGAR DATOS PERSISTENTES
# ----------------------------------------------------
def cargar_datos_persistentes():
    """Carga los datos previamente guardados (incluyendo las rutas de las imágenes)."""
    if os.path.exists(PERSISTENCE_FILE):
        try:
            # Usamos dtype=str para forzar que las columnas de ruta sean siempre strings, evitando el error float
            df = pd.read_csv(PERSISTENCE_FILE, dtype=str) 
            for col in COLUMNAS_IMAGEN:
                if col not in df.columns:
                    df[col] = None 
            return df.to_dict('records')
        except Exception as e:
            st.error(f"Error al cargar datos persistentes: {e}. Se inicia una lista vacía.")
            return []
    return []

# --- 3. Funciones de Lógica y Limpieza ---

def generar_excel_con_formato(df):
    """Genera el archivo Excel en memoria, leyendo las imágenes desde el disco."""
    
    os.makedirs(IMAGE_FOLDER, exist_ok=True) 

    output = BytesIO()
    df_datos_texto = df[ENCABEZADOS] 
    df_datos_texto.to_excel(output, index=False)
    output.seek(0)
    
    wb = load_workbook(output)
    ws = wb.active 
    
    ws.row_dimensions[1].height = 30 
    for row_index in range(1, len(df) + 2):
        fila_excel_obj = ws[row_index]
        if row_index > 1:
            ws.row_dimensions[row_index].height = ALTURA_FILA_PT
            
        for col_idx_0based in range(len(ENCABEZADOS)): 
            celda = fila_excel_obj[col_idx_0based]
            celda.alignment = ALINEACION_CENTRO 
            
            if row_index == 1:
                celda.fill = RELLENO_VERDE_AZULADO
                celda.font = FUENTE_ENCABEZADO
        
    for idx, col_name in enumerate(ENCABEZADOS, start=1):
        col_letter = get_column_letter(idx)
        if col_name == 'OBSERVACIONES':
            ws.column_dimensions[col_letter].width = ANCHO_COLUMNA_OBSERVACIONES_UNITS
        else:
            ws.column_dimensions[col_letter].width = ANCHO_COLUMNA_NORMAL_UNITS

    df_rutas_imagen = df[COLUMNAS_IMAGEN]
    for indice_fila, (index, fila_rutas) in enumerate(df_rutas_imagen.iterrows()):
        fila_excel = indice_fila + 2
        
        for indice_columna, nombre_columna_foto in enumerate(COLUMNAS_IMAGEN):
            
            # 💥 CORRECCIÓN DE ERROR: Usar pd.isna() para verificar si es un valor NaN/float
            imagen_ruta_guardada = fila_rutas[nombre_columna_foto]
            if pd.isna(imagen_ruta_guardada):
                imagen_ruta_guardada = None

            columna_imagen_index = len(ENCABEZADOS) + indice_columna + 1
            columna_letra = get_column_letter(columna_imagen_index)

            if indice_fila == 0: 
                 celda_encabezado = ws[f'{columna_letra}1']
                 celda_encabezado.value = nombre_columna_foto
                 celda_encabezado.alignment = ALINEACION_CENTRO 
                 celda_encabezado.fill = RELLENO_VERDE_AZULADO
                 celda_encabezado.font = FUENTE_ENCABEZADO
                 ws.column_dimensions[columna_letra].width = ANCHO_COLUMNA_NORMAL_UNITS
            
            if imagen_ruta_guardada and os.path.exists(imagen_ruta_guardada):
                try:
                    # Abrir el archivo de disco y forzar la carga completa antes de cerrar el flujo
                    with open(imagen_ruta_guardada, 'rb') as f:
                        img_pil = Image.open(f)
                        img_pil.load() 
                        
                    exif = img_pil._getexif()
                    orientation = 1
                    if exif:
                        for tag, value in exif.items():
                            if TAGS.get(tag, tag) == 'Orientation':
                                orientation = value
                                break
                    
                    if orientation == 3: img_pil = img_pil.rotate(180, expand=True)
                    elif orientation == 6: img_pil = img_pil.rotate(-90, expand=True)
                    elif orientation == 8: img_pil = img_pil.rotate(90, expand=True)

                    rotated_img_bytes = BytesIO()
                    img_pil.save(rotated_img_bytes, format='PNG')
                    rotated_img_bytes.seek(0)

                    img = OpenpyxlImage(rotated_img_bytes)
                    img.width = IMAGEN_WIDTH 
                    img.height = IMAGEN_HEIGHT
                    ws.add_image(img, f'{columna_letra}{fila_excel}')
                except Exception as e:
                    st.warning(f"⚠️ Error al procesar imagen '{imagen_ruta_guardada}': {e}. Insertando sin rotación.")
                    try:
                        # Si falla EXIF/rotación, intentar insertar la imagen original (leyendo los bytes de nuevo)
                        with open(imagen_ruta_guardada, 'rb') as f:
                            img = OpenpyxlImage(f)
                            img.width = IMAGEN_WIDTH 
                            img.height = IMAGEN_HEIGHT
                            ws.add_image(img, f'{columna_letra}{fila_excel}')
                    except Exception as e2:
                         st.error(f"❌ Error crítico al insertar imagen original: {e2}")

    final_output = BytesIO()
    wb.save(final_output)
    return final_output


def guardar_registro_y_limpiar(modelo, serie, condiciones, observaciones, fotos):
    """Guarda el registro, guarda las imágenes físicamente y actualiza el archivo persistente."""
    
    if not modelo or not serie:
        st.error("❌ Los campos MODELO y SERIE son obligatorios.")
        return 

    os.makedirs(IMAGE_FOLDER, exist_ok=True)
    
    # Guardar las fotos físicamente y obtener la RUTA
    fotos_rutas_guardadas = {}
    
    for k, uploaded_file in fotos.items():
        if uploaded_file is not None:
            file_extension = uploaded_file.name.split('.')[-1]
            filename = f"{modelo.replace(' ', '_')}_{serie.replace(' ', '_')}_{k.replace(' ', '_')}.{file_extension}"
            ruta_guardado = os.path.join(IMAGE_FOLDER, filename)
            
            with open(ruta_guardado, "wb") as f:
                f.write(uploaded_file.getbuffer()) 
                
            fotos_rutas_guardadas[k] = ruta_guardado
        else:
            fotos_rutas_guardadas[k] = None

    # 1. Preparar el nuevo registro (con las rutas de las imágenes)
    nuevo_registro = {
        'MODELO': modelo,
        'SERIE': serie,
        'OBSERVACIONES': observaciones,
        **{k: 'SÍ' if v else 'NO' for k, v in condiciones.items()},
        **fotos_rutas_guardadas 
    }
    
    st.session_state['datos_ingresados'].append(nuevo_registro)

    # 2. Persistencia: Guardar el registro a un archivo CSV (texto + rutas)
    try:
        df_actual = pd.DataFrame(st.session_state['datos_ingresados'])
        # Guardamos todas las columnas (texto + rutas) al CSV
        df_actual[COLUMNAS_FINALES].to_csv(PERSISTENCE_FILE, index=False)
    except Exception as e:
        st.warning(f"Advertencia: No se pudo guardar el archivo de persistencia CSV. Error: {e}")
        
    st.success(f"✅ Registro para Modelo {modelo} añadido a la lista. Ingresa el siguiente.")
    
    st.session_state['limpiador_key'] += 1 
    
    keys_a_limpiar_texto = ['input_modelo', 'input_serie', 'input_observaciones']
    for key in keys_a_limpiar_texto:
        if key in st.session_state:
            st.session_state[key] = ""
            
    st.rerun()

# --- 4. Diseño y Tema de la Interfaz Web (Streamlit) ---

st.set_page_config(layout="wide")

LOGO_PATH = "electrolux_logo.png"

# APLICACIÓN DE LOGO Y TÍTULO EN LA BARRA SUPERIOR
try:
    if os.path.exists(LOGO_PATH):
        col_logo, col_titulo = st.columns([1.5, 4]) 
        
        with col_logo:
             st.image(LOGO_PATH, width=150) 
        
        col_titulo.title("Generador de Reporte de Inspección")
        
    else:
        st.title("📋 Generador de Reporte de Inspección")
        st.warning(f"⚠️ ¡Falta el logo! Coloque '{LOGO_PATH}' en esta carpeta.")
except Exception:
    st.title("📋 Generador de Reporte de Inspección")


if 'datos_ingresados' not in st.session_state:
    st.session_state['datos_ingresados'] = cargar_datos_persistentes()
if 'limpiador_key' not in st.session_state:
    st.session_state['limpiador_key'] = 0

# Usamos st.form para agrupar los inputs
with st.form(key='inspeccion_form'):
    
    # TÍTULO EN BLANCO
    st.markdown("<h3 style='color: #FFFFFF;'>--- Datos de Producto y Condición ---</h3>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    
    # Modelo y Serie
    modelo = col1.text_input("MODELO", key='input_modelo')
    serie = col2.text_input("SERIE", key='input_serie')
    
    # Checkboxes
    condiciones = {}
    st.markdown("##### Condiciones de Inspección")
    cols = st.columns(4) 
    for i, col in enumerate([c for c in ENCABEZADOS if c not in ['MODELO', 'SERIE', 'OBSERVACIONES']]):
        clave_dinamica = f'check_{col}_{st.session_state["limpiador_key"]}'
        initial_value = st.session_state.get(clave_dinamica, False)
        condiciones[col] = cols[i % 4].checkbox(col, value=initial_value, key=clave_dinamica)
        
    # Observaciones
    observaciones = st.text_area("OBSERVACIONES", key='input_observaciones')
    
    st.markdown("---")

    # File Uploaders
    st.markdown("<h3 style='color: #FFFFFF;'>📸 Fotos</h3>", unsafe_allow_html=True)
    fotos = {}
    cols_foto = st.columns(4)
    for i, col in enumerate(COLUMNAS_IMAGEN):
        clave_dinamica = f'file_{col}_{st.session_state["limpiador_key"]}'
        fotos[col] = cols_foto[i % 4].file_uploader(col, type=['png', 'jpg', 'jpeg'], key=clave_dinamica)

    st.markdown("---")
    
    # Botón de envío
    st.form_submit_button(
        label="➕ Añadir Registro a la Lista", 
        type="primary", 
        on_click=guardar_registro_y_limpiar,
        args=(modelo, serie, condiciones, observaciones, fotos)
    )

# --- 5. Mostrar tabla de registros guardados y botón de descarga ---

if st.session_state['datos_ingresados']:
    df_preview = pd.DataFrame(st.session_state['datos_ingresados'])
    
    st.subheader(f"Registros Guardados ({len(st.session_state['datos_ingresados'])})")
    st.dataframe(df_preview[ENCABEZADOS], use_container_width=True, height=200)

    excel_file = generar_excel_con_formato(df_preview)
    
    st.download_button(
        label="💾 Descargar Excel Final",
        data=excel_file.getvalue(),
        file_name="Inspeccion_Reporte_Mobil.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if st.button("🗑️ Limpiar Todos los Registros"):
        st.session_state['datos_ingresados'] = []
        st.session_state['limpiador_key'] += 1 
        
        if os.path.exists(PERSISTENCE_FILE):
             os.remove(PERSISTENCE_FILE)
        if os.path.exists(IMAGE_FOLDER):
             shutil.rmtree(IMAGE_FOLDER) 
        
        st.success("Lista de registros y archivos persistentes eliminados.")
        st.rerun()